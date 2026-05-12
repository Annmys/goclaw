#!/usr/bin/env python3
"""Golden-file regression for shipping document generation.

This script never copies the completed workbook as output. It runs the active
generate_shipping_doc.py script against "*初始.xlsx" inputs and compares the
generated workbook with the matching "*完成.xlsx" golden workbook.
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_CASE_DIRS = [
    Path("/mnt/test-data/船务清单"),
    Path(r"D:\数据\测试数据\船务清单"),
]
DEFAULT_OUTPUT_ROOTS = [
    Path("/app/workspace/system/evolution-regression/shipping-doc-processing"),
    Path(r"D:\数据\存储\system\evolution-regression\shipping-doc-processing"),
]


@dataclass
class GoldenCase:
    name: str
    input_file: Path
    golden_file: Path


def existing_default_case_dir() -> Path | None:
    for path in DEFAULT_CASE_DIRS:
        if path.exists() and path.is_dir():
            return path
    return None


def existing_default_output_root() -> Path:
    for path in DEFAULT_OUTPUT_ROOTS:
        parent = path.parent
        if parent.exists() or str(path).startswith("/app/"):
            return path
    return Path(tempfile.gettempdir()) / "goclaw-shipping-golden"


def discover_cases(cases_dir: Path, names: list[str] | None = None) -> list[GoldenCase]:
    cases: list[GoldenCase] = []
    filters = set(names or [])
    for input_file in sorted(cases_dir.glob("*初始.xlsx")):
        name = input_file.name.replace("-初始.xlsx", "")
        if filters and name not in filters:
            continue
        golden_file = cases_dir / f"{name}-完成.xlsx"
        if not golden_file.exists():
            continue
        cases.append(GoldenCase(name=name, input_file=input_file, golden_file=golden_file))
    return cases


def find_sheet(workbook, names: tuple[str, ...]):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    for sheet in workbook.worksheets:
        if sheet.title.upper() in {name.upper() for name in names}:
            return sheet
    return None


def normalized(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\u00a0", " ")


def compact(value: Any) -> str:
    return " ".join(normalized(value).split()).upper()


def non_empty_count(ws, min_row: int = 1, max_row: int | None = None) -> int:
    total = 0
    max_row = max_row or ws.max_row
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        for cell in row:
            if normalized(cell.value):
                total += 1
    return total


def image_count(ws) -> int:
    return len(getattr(ws, "_images", []) or [])


def merged_ranges(ws) -> set[str]:
    return {str(item) for item in ws.merged_cells.ranges}


def column_widths(ws) -> dict[str, float]:
    widths: dict[str, float] = {}
    for key, dim in ws.column_dimensions.items():
        if dim.width:
            widths[key] = float(dim.width)
    return widths


def find_cell_contains(ws, keywords: tuple[str, ...], max_row: int = 40) -> tuple[int, int] | None:
    wanted = tuple(keyword.upper() for keyword in keywords)
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row)):
        for cell in row:
            text = compact(cell.value)
            if text and all(keyword in text for keyword in wanted):
                return cell.row, cell.column
    return None


def find_header_col(ws, labels: tuple[str, ...], max_row: int = 40) -> int | None:
    wanted = tuple(label.upper() for label in labels)
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row)):
        for cell in row:
            text = compact(cell.value)
            if not text:
                continue
            if any(label in text for label in wanted):
                return cell.column
    return None


def find_total_row(ws) -> int | None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            text = compact(cell.value)
            if text in {"TOTAL", "TOTAL:", "合计"} or text.startswith("TOTAL "):
                return cell.row
    return None


def detail_range(ws) -> tuple[int, int] | None:
    header = find_cell_contains(ws, ("NO",), max_row=40)
    total_row = find_total_row(ws)
    if not header or not total_row or total_row <= header[0] + 1:
        return None
    return header[0] + 1, total_row - 1


def values_in_col(ws, col: int | None, row_range: tuple[int, int] | None) -> list[str]:
    if not col or not row_range:
        return []
    start, end = row_range
    values: list[str] = []
    for row in range(start, end + 1):
        value = normalized(ws.cell(row=row, column=col).value)
        if value:
            values.append(value)
    return values


def add_score(checks: list[dict[str, Any]], name: str, passed: bool, points: int, message: str) -> int:
    checks.append({"name": name, "passed": passed, "points": points if passed else 0, "max_points": points, "message": message})
    return points if passed else 0


def compare_workbooks(generated: Path, golden: Path) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []
    score = 0
    max_score = 100
    failures: list[str] = []

    gen_wb = load_workbook(generated)
    gold_wb = load_workbook(golden)
    try:
        gen_ci = find_sheet(gen_wb, ("CI",))
        gold_ci = find_sheet(gold_wb, ("CI",))
        gen_epl = find_sheet(gen_wb, ("EPL", "PL"))
        gold_epl = find_sheet(gold_wb, ("EPL", "PL"))

        sheet_ok = gen_ci is not None and gen_epl is not None
        score += add_score(checks, "required_sheets", sheet_ok, 10, f"generated sheets={gen_wb.sheetnames}")
        if not sheet_ok:
            failures.append("missing CI or EPL/PL sheet")
            return {"score": score, "max_score": max_score, "checks": checks, "failures": failures}

        assert gen_ci is not None and gen_epl is not None

        if gold_ci is not None:
            gen_ci_non_empty = non_empty_count(gen_ci, max_row=min(40, gen_ci.max_row))
            gold_ci_non_empty = non_empty_count(gold_ci, max_row=min(40, gold_ci.max_row))
            ci_text_ok = gen_ci_non_empty >= max(1, int(gold_ci_non_empty * 0.9))
            score += add_score(checks, "ci_text_preserved", ci_text_ok, 10, f"CI non-empty {gen_ci_non_empty}/{gold_ci_non_empty}")
            if not ci_text_ok:
                failures.append("CI text content was not sufficiently preserved")

            ci_merge_ok = len(merged_ranges(gen_ci)) >= max(0, int(len(merged_ranges(gold_ci)) * 0.8))
            score += add_score(checks, "ci_merges_preserved", ci_merge_ok, 8, f"CI merges {len(merged_ranges(gen_ci))}/{len(merged_ranges(gold_ci))}")
            if not ci_merge_ok:
                failures.append("CI merged-cell layout dropped too much")

            ci_width_ok = len(column_widths(gen_ci)) >= max(0, int(len(column_widths(gold_ci)) * 0.8))
            score += add_score(checks, "ci_widths_preserved", ci_width_ok, 6, f"CI widths {len(column_widths(gen_ci))}/{len(column_widths(gold_ci))}")
            if not ci_width_ok:
                failures.append("CI column widths dropped too much")

            gold_images = image_count(gold_ci)
            if gold_images:
                image_ok = image_count(gen_ci) >= gold_images
                score += add_score(checks, "ci_images_preserved", image_ok, 6, f"CI images {image_count(gen_ci)}/{gold_images}")
                if not image_ok:
                    failures.append("CI images/logo were not preserved")
            else:
                score += add_score(checks, "ci_images_preserved", True, 6, "golden has no CI images")
        else:
            score += add_score(checks, "ci_reference_available", True, 30, "golden has no CI sheet; skipped CI layout comparison")

        if gold_epl is None:
            score += add_score(checks, "epl_reference_available", False, 20, "golden has no EPL/PL sheet")
            failures.append("golden missing EPL/PL sheet")
            return {"score": score, "max_score": max_score, "checks": checks, "failures": failures}

        gen_range = detail_range(gen_epl)
        gold_range = detail_range(gold_epl)
        range_ok = gen_range is not None and gold_range is not None
        score += add_score(checks, "epl_detail_range_detected", range_ok, 8, f"generated={gen_range}, golden={gold_range}")
        if not range_ok:
            failures.append("EPL detail range or Total row was not detectable")

        if gen_range and gold_range:
            gen_rows = gen_range[1] - gen_range[0] + 1
            gold_rows = gold_range[1] - gold_range[0] + 1
            row_ok = abs(gen_rows - gold_rows) <= max(2, int(gold_rows * 0.25))
            score += add_score(checks, "epl_detail_row_count", row_ok, 8, f"EPL detail rows {gen_rows}/{gold_rows}")
            if not row_ok:
                failures.append("EPL detail row count differs too much from golden")

            after_total = non_empty_count(gen_epl, min_row=(find_total_row(gen_epl) or gen_range[1]) + 1, max_row=min(gen_epl.max_row, (find_total_row(gen_epl) or gen_range[1]) + 20))
            no_garbage_ok = after_total <= 4
            score += add_score(checks, "no_garbage_after_total", no_garbage_ok, 8, f"non-empty cells after Total={after_total}")
            if not no_garbage_ok:
                failures.append("generated EPL has too much content after Total")

        required_cols = {
            "item_no": ("ITEM NO", "ITEM"),
            "qty": ("QTY",),
            "c_n": ("C/N",),
            "gross_weight": ("G.W", "G.W."),
            "dimension": ("DIMENSION",),
        }
        found_cols = {name: find_header_col(gen_epl, labels) for name, labels in required_cols.items()}
        cols_ok = all(found_cols.values())
        score += add_score(checks, "epl_required_columns", cols_ok, 12, f"columns={found_cols}")
        if not cols_ok:
            failures.append("EPL required columns are missing")

        gold_cols = {name: find_header_col(gold_epl, labels) for name, labels in required_cols.items()}
        for col_name, points in [("c_n", 6), ("gross_weight", 6), ("dimension", 8)]:
            gen_values = values_in_col(gen_epl, found_cols.get(col_name), gen_range)
            gold_values = values_in_col(gold_epl, gold_cols.get(col_name), gold_range)
            if gold_values:
                ok = bool(gen_values)
                score += add_score(checks, f"epl_{col_name}_filled", ok, points, f"{col_name}: generated={len(gen_values)}, golden={len(gold_values)}")
                if not ok:
                    failures.append(f"EPL {col_name} is empty")
            else:
                score += add_score(checks, f"epl_{col_name}_filled", True, points, f"golden has no {col_name} values")

        gen_dim_values = values_in_col(gen_epl, found_cols.get("dimension"), gen_range)
        gold_dim_values = values_in_col(gold_epl, gold_cols.get("dimension"), gold_range)
        gold_dim_lines = sum(value.count("\n") + 1 for value in gold_dim_values if "*" in value)
        gen_dim_lines = sum(value.count("\n") + 1 for value in gen_dim_values if "*" in value)
        multi_pack_ok = gold_dim_lines <= 1 or gen_dim_lines >= max(1, int(gold_dim_lines * 0.8))
        score += add_score(checks, "multi_packaging_lines_preserved", multi_pack_ok, 4, f"dimension lines {gen_dim_lines}/{gold_dim_lines}")
        if not multi_pack_ok:
            failures.append("multi-line packaging appears collapsed")

        gen_merge_count = len(merged_ranges(gen_epl))
        gold_merge_count = len(merged_ranges(gold_epl))
        merge_ok = gen_merge_count >= max(0, int(gold_merge_count * 0.65))
        score += add_score(checks, "epl_merges_preserved", merge_ok, 6, f"EPL merges {gen_merge_count}/{gold_merge_count}")
        if not merge_ok:
            failures.append("EPL merged-cell layout dropped too much")

    finally:
        gen_wb.close()
        gold_wb.close()

    return {
        "score": score,
        "max_score": max_score,
        "checks": checks,
        "failures": failures,
    }


def run_generation(input_file: Path, output_file: Path) -> dict[str, Any]:
    script = SCRIPT_DIR / "generate_shipping_doc.py"
    proc = subprocess.run(
        [sys.executable, str(script), str(input_file), str(output_file)],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=180,
    )
    result: dict[str, Any] = {
        "returncode": proc.returncode,
        "stdout": proc.stdout,
        "stderr": proc.stderr,
    }
    try:
        result["json"] = json.loads(proc.stdout)
    except Exception:
        result["json"] = None
    return result


def run_case(case: GoldenCase, output_dir: Path, min_score: int) -> dict[str, Any]:
    generated = output_dir / f"{case.name}-golden-regression-output.xlsx"
    generation = run_generation(case.input_file, generated)
    if generation["returncode"] != 0 or not generated.exists():
        return {
            "name": case.name,
            "status": "failed",
            "score": 0,
            "min_score": min_score,
            "input_file": str(case.input_file),
            "golden_file": str(case.golden_file),
            "generated_file": str(generated),
            "failures": ["generation failed"],
            "generation": generation,
        }

    comparison = compare_workbooks(generated, case.golden_file)
    status = "passed" if comparison["score"] >= min_score and not comparison["failures"] else "failed"
    return {
        "name": case.name,
        "status": status,
        "score": comparison["score"],
        "min_score": min_score,
        "input_file": str(case.input_file),
        "golden_file": str(case.golden_file),
        "generated_file": str(generated),
        "failures": comparison["failures"],
        "checks": comparison["checks"],
        "generation": generation.get("json") or {"stdout": generation["stdout"], "stderr": generation["stderr"]},
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--cases-dir", type=Path, default=existing_default_case_dir())
    parser.add_argument("--output-dir", type=Path, default=existing_default_output_root())
    parser.add_argument("--case", action="append", dest="case_names", help="Case name, for example 测试订单7")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--min-score", type=int, default=75)
    args = parser.parse_args()

    if args.cases_dir is None or not args.cases_dir.exists():
        print(json.dumps({"ok": False, "error": "cases directory not found"}, ensure_ascii=False))
        return 1

    output_dir: Path = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    cases = discover_cases(args.cases_dir, args.case_names)
    if args.limit and args.limit > 0:
        cases = cases[: args.limit]
    if not cases:
        print(json.dumps({"ok": False, "error": "no golden cases found", "cases_dir": str(args.cases_dir)}, ensure_ascii=False))
        return 1

    results = [run_case(case, output_dir, args.min_score) for case in cases]
    passed = sum(1 for item in results if item["status"] == "passed")
    failed = sum(1 for item in results if item["status"] == "failed")
    payload = {
        "ok": failed == 0,
        "cases_dir": str(args.cases_dir),
        "output_dir": str(output_dir),
        "total": len(results),
        "passed": passed,
        "failed": failed,
        "cases": results,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
