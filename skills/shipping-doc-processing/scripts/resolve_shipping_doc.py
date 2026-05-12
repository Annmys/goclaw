#!/usr/bin/env python3
import json
import re
import sqlite3
import sys
from pathlib import Path

try:
    import xlrd
except Exception:
    xlrd = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


FLOW_ROOT_CANDIDATES = [
    Path("/mnt/target/flow-orders"),
    Path(r"D:\数据\包装流转单"),
]


def is_valid_sqlite_table(path: Path, table_name: str) -> bool:
    try:
        if not path.exists() or not path.is_file() or path.stat().st_size <= 0:
            return False
        conn = sqlite3.connect(path)
        try:
            cur = conn.cursor()
            cur.execute(
                "select name from sqlite_master where type='table' and name=?",
                (table_name,),
            )
            return cur.fetchone() is not None
        finally:
            conn.close()
    except Exception:
        return False


def find_sqlite(table_name: str) -> Path | None:
    candidates: list[Path] = []
    for root in FLOW_ROOT_CANDIDATES:
        if root.exists():
            candidates.extend(root.glob("*.sqlite"))
    for candidate in candidates:
        if is_valid_sqlite_table(candidate, table_name):
            return candidate
    return None


def read_workbook_preview(path: Path) -> tuple[list[str], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        if load_workbook is None:
            raise RuntimeError("openpyxl not available")
        workbook = load_workbook(path, read_only=True, data_only=True)
        values: list[str] = []
        try:
            sheets = workbook.sheetnames[:]
            for sheet in workbook.worksheets[:3]:
                for row in sheet.iter_rows(min_row=1, max_row=20, max_col=20, values_only=True):
                    values.extend(str(cell).strip() for cell in row if cell not in (None, ""))
        finally:
            workbook.close()
        return sheets, values

    if suffix == ".xls":
        if xlrd is None:
            raise RuntimeError("xlrd not available")
        workbook = xlrd.open_workbook(str(path), on_demand=True)
        values: list[str] = []
        sheets = workbook.sheet_names()
        try:
            for name in sheets[:3]:
                sheet = workbook.sheet_by_name(name)
                for row in range(min(sheet.nrows, 20)):
                    for col in range(min(sheet.ncols, 20)):
                        cell = sheet.cell_value(row, col)
                        if cell not in ("", None):
                            values.append(str(cell).strip())
        finally:
            workbook.release_resources()
        return sheets, values

    raise RuntimeError(f"unsupported workbook type: {suffix}")


def extract_xs_order_no(path: Path, values: list[str]) -> str | None:
    match = re.search(r"(XS\d+)", path.name, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    for value in values:
        match = re.search(r"(XS\d+)", value, re.IGNORECASE)
        if match:
            return match.group(1).upper()
    return None


def inspect_shipping_completeness(sheet_names: list[str], values: list[str]) -> dict:
    lower_names = [name.lower() for name in sheet_names]
    combined = "\n".join(values).lower()
    has_ci_sheet = any(name in lower_names for name in ["ci", "invoice"])
    has_pl_sheet = any(name in lower_names for name in ["pl", "epl", "packing list"])
    has_invoice_keywords = "commercial invoice" in combined
    has_packing_keywords = "packing list" in combined or "g.w.(kg)" in combined or "dimension" in combined

    missing = []
    if not has_ci_sheet:
        missing.append("缺少 CI sheet")
    if not has_pl_sheet:
        missing.append("缺少 PL/EPL sheet")

    return {
        "has_ci_sheet": has_ci_sheet,
        "has_pl_sheet": has_pl_sheet,
        "has_invoice_keywords": has_invoice_keywords,
        "has_packing_keywords": has_packing_keywords,
        "is_complete": has_ci_sheet and has_pl_sheet,
        "missing_items": missing,
    }


def lookup_content_index(xs_order_no: str) -> dict | None:
    content_db = find_sqlite("flow_content_index")
    if content_db is None:
        return None

    conn = sqlite3.connect(content_db)
    try:
        cur = conn.cursor()
        cur.execute(
            """
            select year_folder, owner_folder, workbook_name, sheet_name, workbook_path,
                   outer_box_row, outer_box_spec, outer_box_qty, dimension_for_epl, c_n_for_epl
            from flow_content_index
            where upper(order_no) = upper(?)
            limit 1
            """,
            (xs_order_no,),
        )
        row = cur.fetchone()
        if not row:
            return None
        return {
            "content_db": str(content_db),
            "year_folder": row[0],
            "owner_folder": row[1],
            "workbook_name": row[2],
            "sheet_name": row[3],
            "workbook_path": row[4],
            "outer_box_row": row[5],
            "outer_box_spec": row[6],
            "outer_box_qty": row[7],
            "dimension_for_epl": row[8],
            "c_n_for_epl": row[9],
        }
    finally:
        conn.close()


def lookup_order_mapping(xs_order_no: str) -> dict | None:
    mapping_db = find_sqlite("order_mapping")
    if mapping_db is None:
        return None

    conn = sqlite3.connect(mapping_db)
    try:
        cur = conn.cursor()
        cur.execute(
            """
            select year_folder, owner_folder, workbook_name, sheet_name, workbook_path
            from order_mapping
            where upper(order_no) = upper(?)
            limit 1
            """,
            (xs_order_no,),
        )
        row = cur.fetchone()
        if not row:
            return None
        return {
            "mapping_db": str(mapping_db),
            "year_folder": row[0],
            "owner_folder": row[1],
            "workbook_name": row[2],
            "sheet_name": row[3],
            "workbook_path": row[4],
        }
    finally:
        conn.close()


def looks_like_dimension(text: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.\d+)?\s*\*\s*\d+(?:\.\d+)?\s*\*\s*\d+(?:\.\d+)?", text))


def normalize_number(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return str(number).rstrip("0").rstrip(".")


def extract_primary_package_boxes(rows) -> list[dict]:
    boxes: list[dict] = []
    for row_index, row in enumerate(rows, start=1):
        values = [str(value).strip() if value not in (None, "") else "" for value in row]
        if len(values) < 5:
            continue
        label0 = values[0]
        label1 = values[1]
        spec = values[3]
        qty = normalize_number(values[4])
        if label1 not in {"外箱", "通口箱", "啤盒6.5", "啤盒7.5"}:
            continue
        if not looks_like_dimension(spec):
            continue
        if not qty or not re.fullmatch(r"\d+(?:\.\d+)?", qty):
            continue
        boxes.append(
            {
                "row": row_index,
                "outer_box_spec": spec.replace(" ", ""),
                "outer_box_qty": qty,
                "dimension_for_epl": f"{spec.replace(' ', '')}CM*{qty}",
                "label": label1,
                "group": label0,
            }
        )
    return boxes


def extract_outer_box(rows) -> dict:
    primary_boxes = extract_primary_package_boxes(rows)
    if primary_boxes:
        carton_total = sum(int(float(box["outer_box_qty"])) for box in primary_boxes)
        dimension_text = "\n".join(box["dimension_for_epl"] for box in primary_boxes)
        return {
            "row": primary_boxes[0]["row"],
            "outer_box_spec": primary_boxes[0]["outer_box_spec"],
            "outer_box_qty": str(carton_total),
            "dimension_for_epl": dimension_text,
            "c_n_for_epl": f"1-{carton_total}",
        }

    for row_index, row in enumerate(rows, start=1):
        values = [str(value).strip() if value not in (None, "") else "" for value in row]
        for idx, value in enumerate(values):
            if not looks_like_dimension(value):
                continue
            label_window = " ".join(values[max(0, idx - 3) : idx + 1])
            if "箱" not in label_window and "CTN" not in label_window.upper():
                continue
            qty = ""
            for next_value in values[idx + 1 : idx + 4]:
                candidate = normalize_number(next_value)
                if candidate and re.fullmatch(r"\d+(?:\.\d+)?", candidate):
                    qty = candidate
                    break
            if not qty:
                continue
            carton_count = int(float(qty))
            return {
                "row": row_index,
                "outer_box_spec": value.replace(" ", ""),
                "outer_box_qty": qty,
                "dimension_for_epl": f"{value.replace(' ', '')}CM*{qty}",
                "c_n_for_epl": f"1-{carton_count}",
            }
    raise RuntimeError("outer box row not found")


def read_flow_order_outer_box(flow_path: Path, sheet_name: str) -> dict:
    suffix = flow_path.suffix.lower()
    if suffix == ".xlsx":
        if load_workbook is None:
            raise RuntimeError("openpyxl not available")
        workbook = load_workbook(flow_path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            return extract_outer_box(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

    if suffix == ".xls":
        if xlrd is None:
            raise RuntimeError("xlrd not available")
        workbook = xlrd.open_workbook(str(flow_path), on_demand=True)
        try:
            sheet = workbook.sheet_by_name(sheet_name) if sheet_name in workbook.sheet_names() else workbook.sheet_by_index(0)
            rows = ([sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(sheet.nrows))
            return extract_outer_box(rows)
        finally:
            workbook.release_resources()

    raise RuntimeError(f"unsupported flow order file: {flow_path}")


def resolve_shipping_doc(shipping_file: Path) -> dict:
    sheet_names, values = read_workbook_preview(shipping_file)
    xs_order_no = extract_xs_order_no(shipping_file, values)
    completeness = inspect_shipping_completeness(sheet_names, values)
    content_hit = lookup_content_index(xs_order_no) if xs_order_no else None
    mapping = lookup_order_mapping(xs_order_no) if xs_order_no and not content_hit else None

    result = {
        "shipping_file": str(shipping_file),
        "sheet_names": sheet_names,
        "xs_order_no": xs_order_no,
        "shipping_completeness": completeness,
        "content_index": content_hit,
        "mapping": mapping,
    }

    if content_hit:
        result["outer_box"] = {
            "row": content_hit.get("outer_box_row"),
            "outer_box_spec": content_hit.get("outer_box_spec"),
            "outer_box_qty": content_hit.get("outer_box_qty"),
            "dimension_for_epl": content_hit.get("dimension_for_epl"),
            "c_n_for_epl": content_hit.get("c_n_for_epl"),
        }
        result["epl_fill_suggestion"] = {
            "packaging_source": "流转单内容索引",
            "dimension": content_hit.get("dimension_for_epl"),
            "c_n": content_hit.get("c_n_for_epl"),
            "carton_count": content_hit.get("outer_box_qty"),
        }
        return result

    if mapping:
        outer_box = read_flow_order_outer_box(Path(mapping["workbook_path"]), str(mapping["sheet_name"]))
        result["outer_box"] = outer_box
        result["epl_fill_suggestion"] = {
            "packaging_source": "原始流转单",
            "dimension": outer_box.get("dimension_for_epl"),
            "c_n": outer_box.get("c_n_for_epl"),
            "carton_count": outer_box.get("outer_box_qty"),
        }
        return result

    result["outer_box"] = None
    result["epl_fill_suggestion"] = None
    return result


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: resolve_shipping_doc.py <shipping-file>", file=sys.stderr)
        return 2

    shipping_file = Path(sys.argv[1])
    if not shipping_file.exists():
        print(json.dumps({"error": f"file not found: {shipping_file}"}, ensure_ascii=False))
        return 1

    try:
        result = resolve_shipping_doc(shipping_file)
    except Exception as exc:
        print(json.dumps({"shipping_file": str(shipping_file), "error": str(exc)}, ensure_ascii=False))
        return 1

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
