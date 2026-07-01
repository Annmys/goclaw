#!/usr/bin/env python3
import csv
import json
import re
import sys
from pathlib import Path

try:
    import xlrd
except Exception:
    xlrd = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


RULES = [
    {
        "type": "流转单",
        "keywords": ["流转单", "包装材料需求流转单", "订单号", "外箱"],
        "next_actions": ["提取订单号", "定位外箱信息", "必要时进入流转单查询流程"],
        "handoff_skill": "",
    },
    {
        "type": "订单映射表",
        "keywords": ["订单映射", "年份目录", "workbook_path", "sheet_name"],
        "next_actions": ["按订单号查询", "查重", "筛选异常记录"],
        "handoff_skill": "",
    },
    {
        "type": "装箱单/PL/EPL",
        "keywords": ["packing list", "estimated packing list", "c/n", "dimension", "cbm"],
        "next_actions": ["核对箱数和尺寸", "判断是否属于预估箱单制作范围"],
        "handoff_skill": "",
    },
    {
        "type": "商业发票/CI",
        "keywords": ["commercial invoice", "invoice", "amount", "usd"],
        "next_actions": ["提取金额和明细", "判断是否属于预估箱单制作范围"],
        "handoff_skill": "",
    },
    {
        "type": "采购单/PO",
        "keywords": ["purchase order", "采购订单", "po"],
        "next_actions": ["提取采购信息", "核对供应商和交期"],
        "handoff_skill": "",
    },
]


def read_csv_preview(path: Path) -> tuple[list[str], list[str]]:
    values: list[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for idx, row in enumerate(reader):
            values.extend(str(cell).strip() for cell in row if str(cell).strip())
            if idx >= 9:
                break
    return ["CSV"], values


def read_xlsx_preview(path: Path) -> tuple[list[str], list[str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl not available")
    workbook = load_workbook(path, read_only=True, data_only=True)
    values: list[str] = []
    try:
        sheets = workbook.sheetnames[:]
        for sheet in workbook.worksheets[:3]:
            for row in sheet.iter_rows(min_row=1, max_row=20, max_col=20, values_only=True):
                values.extend(str(cell).strip() for cell in row if cell not in (None, ""))
    finally:
        workbook.close()
    return sheets, values


def read_xls_preview(path: Path) -> tuple[list[str], list[str]]:
    if xlrd is None:
        raise RuntimeError("xlrd not available")
    workbook = xlrd.open_workbook(str(path), on_demand=True)
    values: list[str] = []
    sheets = workbook.sheet_names()
    try:
        for name in sheets[:3]:
            sheet = workbook.sheet_by_name(name)
            for row in range(min(sheet.nrows, 20)):
                for col in range(min(sheet.ncols, 20)):
                    cell = sheet.cell_value(row, col)
                    if cell not in ("", None):
                        values.append(str(cell).strip())
    finally:
        workbook.release_resources()
    return sheets, values


def read_preview(path: Path) -> tuple[list[str], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_preview(path)
    if suffix == ".xlsx":
        return read_xlsx_preview(path)
    if suffix == ".xls":
        return read_xls_preview(path)
    raise RuntimeError(f"unsupported file type: {suffix}")


def extract_xs_order_no(path_name: str, values: list[str]) -> str | None:
    match = re.search(r"(XS\d+)", path_name, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    for value in values:
        match = re.search(r"(XS\d+)", value, re.IGNORECASE)
        if match:
            return match.group(1).upper()
    return None


def detect_shipping_doc(sheet_names: list[str], values: list[str]) -> tuple[bool, dict]:
    lower_names = [name.lower() for name in sheet_names]
    combined = "\n".join(values).lower()
    has_ci_sheet = any(name in lower_names for name in ["ci", "invoice"])
    has_pl_sheet = any(name in lower_names for name in ["pl", "epl", "packing list"])
    has_invoice_keywords = "commercial invoice" in combined
    has_packing_keywords = "packing list" in combined or "g.w.(kg)" in combined or "dimension" in combined

    missing = []
    if not has_ci_sheet:
        missing.append("缺少 CI sheet")
    if not has_pl_sheet:
        missing.append("缺少 PL/EPL sheet")

    detail = {
        "has_ci_sheet": has_ci_sheet,
        "has_pl_sheet": has_pl_sheet,
        "has_invoice_keywords": has_invoice_keywords,
        "has_packing_keywords": has_packing_keywords,
        "missing_items": missing,
        "handoff_skill": "epl-core-workflow（预估箱单制作核心流程）V2",
    }
    matched = (has_ci_sheet and has_pl_sheet) or (has_invoice_keywords and has_packing_keywords)
    return matched, detail


def detect_type(path: Path) -> dict:
    sheet_names, values = read_preview(path)
    xs_order_no = extract_xs_order_no(path.name, values)
    combined = "\n".join([path.name, *sheet_names, *values]).lower()

    shipping_hit, shipping_detail = detect_shipping_doc(sheet_names, values)
    if shipping_hit:
        return {
            "file": str(path),
            "detected_type": "船务清单",
            "confidence": "high" if shipping_detail["has_ci_sheet"] and shipping_detail["has_pl_sheet"] else "medium",
            "sheet_names": sheet_names,
            "xs_order_no": xs_order_no,
            "matched_rules": ["具备船务清单典型 sheet 或关键词特征"],
            "next_actions": [
                "调用核心 skill epl-core-workflow（预估箱单制作核心流程）V2",
                "检查 CI 和 EPL/PL 是否完整",
                "如需补全，由epl-core-workflow（预估箱单制作核心流程）V2 skill 负责后续流转单查询和包装补全",
            ],
            "handoff_skill": "epl-core-workflow（预估箱单制作核心流程）V2",
            "shipping_summary": shipping_detail,
        }

    if xs_order_no and (("commercial invoice" in combined) or any(name.lower() in ["ci", "invoice"] for name in sheet_names)):
        return {
            "file": str(path),
            "detected_type": "船务清单",
            "confidence": "medium",
            "sheet_names": sheet_names,
            "xs_order_no": xs_order_no,
            "matched_rules": ["CI-only 文件包含 XS 订单号，判定为待补全船务清单"],
            "next_actions": [
                "调用核心 skill epl-core-workflow（预估箱单制作核心流程）V2",
                "补齐缺少的 EPL/PL",
                "根据流转单和产品包装重量表生成完成版船务清单",
            ],
            "handoff_skill": "epl-core-workflow（预估箱单制作核心流程）V2",
            "shipping_summary": {
                "has_ci_sheet": any(name.lower() in ["ci", "invoice"] for name in sheet_names),
                "has_pl_sheet": any(name.lower() in ["pl", "epl", "packing list"] for name in sheet_names),
                "has_invoice_keywords": "commercial invoice" in combined,
                "has_packing_keywords": "packing list" in combined or "g.w.(kg)" in combined or "dimension" in combined,
                "missing_items": ["缺少 PL/EPL sheet"],
                "handoff_skill": "epl-core-workflow（预估箱单制作核心流程）V2",
            },
        }

    scored: list[tuple[int, dict, list[str]]] = []
    for rule in RULES:
        matched = [keyword for keyword in rule["keywords"] if keyword.lower() in combined]
        if matched:
            scored.append((len(matched), rule, matched))

    scored.sort(key=lambda item: item[0], reverse=True)
    if not scored:
        return {
            "file": str(path),
            "detected_type": "unknown",
            "confidence": "low",
            "sheet_names": sheet_names,
            "xs_order_no": xs_order_no,
            "matched_rules": [],
            "next_actions": ["查看文件名", "查看 sheet 名称", "查看前 10 行表头后再决定流程"],
            "handoff_skill": "",
        }

    score, rule, matched = scored[0]
    return {
        "file": str(path),
        "detected_type": rule["type"],
        "confidence": "high" if score >= 3 else "medium",
        "sheet_names": sheet_names,
        "xs_order_no": xs_order_no,
        "matched_rules": matched,
        "next_actions": rule["next_actions"],
        "handoff_skill": rule["handoff_skill"],
    }


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: detect_excel_type.py <excel-file>", file=sys.stderr)
        return 2

    path = Path(sys.argv[1])
    if not path.exists():
        print(json.dumps({"error": f"file not found: {path}"}, ensure_ascii=False))
        return 1

    try:
        result = detect_type(path)
    except Exception as exc:
        print(json.dumps({"file": str(path), "error": str(exc)}, ensure_ascii=False))
        return 1

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
