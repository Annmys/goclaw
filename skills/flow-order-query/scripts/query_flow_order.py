#!/usr/bin/env python3
import json
import re
import sqlite3
import sys
from pathlib import Path

try:
    import xlrd
except Exception:
    xlrd = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

ROOT_CANDIDATES = [
    Path("/mnt/target/flow-orders"),
    Path(r"D:\数据\包装流转单"),
]


def is_valid_db(path: Path, table_name: str) -> bool:
    try:
        if not path.exists() or not path.is_file() or path.stat().st_size <= 0:
            return False
        conn = sqlite3.connect(path)
        try:
            cur = conn.cursor()
            cur.execute("select name from sqlite_master where type='table' and name=?", (table_name,))
            return cur.fetchone() is not None
        finally:
            conn.close()
    except Exception:
        return False


def find_content_db() -> Path | None:
    candidates = []
    for root in ROOT_CANDIDATES:
        if not root.exists():
            continue
        candidates.extend(root.glob("*.sqlite"))

    preferred = [p for p in candidates if "流转单内容索引" in p.name]
    for path in [*preferred, *candidates]:
        if is_valid_db(path, "flow_content_index"):
            return path
    return None


def find_mapping_db() -> Path | None:
    candidates = []
    for root in ROOT_CANDIDATES:
        if not root.exists():
            continue
        candidates.extend(root.glob("*.sqlite"))

    preferred = [p for p in candidates if "订单映射表" in p.name]
    for path in [*preferred, *candidates]:
        if is_valid_db(path, "order_mapping"):
            return path
    return None


def extract_order_no(text: str) -> str | None:
    match = re.search(r"(XS\d+)", text, re.IGNORECASE)
    return match.group(1).upper() if match else None


def lookup_mapping(order_no: str) -> dict | None:
    db = find_mapping_db()
    if db is None:
        return None

    conn = sqlite3.connect(db)
    try:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute(
            """
            select order_no, year_folder, owner_folder, workbook_name, sheet_name, workbook_path, status
            from order_mapping
            where upper(order_no)=upper(?)
            order by workbook_name desc
            limit 1
            """,
            (order_no,),
        )
        row = cur.fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def extract_outer_box(rows) -> dict | None:
    for row_index, row in enumerate(rows, start=1):
        vals = [str(v).strip() if v not in (None, "") else "" for v in row]
        label_idx = -1
        for idx, value in enumerate(vals):
            if value == "外箱" or value in {"通口箱", "纸箱", "外纸箱"}:
                label_idx = idx
                break
            if value.endswith("箱") and value not in {"箱", "木箱"}:
                label_idx = idx
                break
        if label_idx < 0:
            continue
        idx = label_idx
        if vals[idx] == "外箱":
            spec = vals[idx + 2] if idx + 2 < len(vals) else ""
            qty = vals[idx + 3] if idx + 3 < len(vals) else ""
        else:
            spec = vals[3] if len(vals) > 3 else ""
            qty = vals[4] if len(vals) > 4 else ""
        qty_text = str(qty).strip()
        c_n = ""
        if qty_text:
            try:
                c_n = f"1-{int(float(qty_text))}"
            except Exception:
                c_n = ""
        dimension = f"{spec}CM*{qty_text}" if spec and qty_text else ""
        return {
            "outer_box_row": str(row_index),
            "outer_box_label": vals[idx],
            "outer_box_spec": spec,
            "outer_box_qty": qty_text,
            "dimension_for_epl": dimension,
            "c_n_for_epl": c_n,
            "source": "flow_outer_box_fallback",
        }
    return None


def read_outer_box_from_workbook(workbook_path: str, sheet_name: str) -> dict | None:
    path = Path(workbook_path)
    if not path.exists():
        return None

    if path.suffix.lower() == ".xlsx":
        if load_workbook is None:
            return None
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            return extract_outer_box(ws.iter_rows(values_only=True))
        finally:
            wb.close()

    if path.suffix.lower() == ".xls":
        if xlrd is None:
            return None
        wb = xlrd.open_workbook(str(path), on_demand=True)
        try:
            ws = wb.sheet_by_name(sheet_name) if sheet_name in wb.sheet_names() else wb.sheet_by_index(0)
            rows = ([ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows))
            return extract_outer_box(rows)
        finally:
            wb.release_resources()

    return None


def query_order(order_no: str) -> dict:
    db = find_content_db()
    if db is None:
        return {
            "ok": False,
            "order_no": order_no,
            "error": "flow content index sqlite not found",
            "searched_roots": [str(p) for p in ROOT_CANDIDATES],
        }

    conn = sqlite3.connect(db)
    try:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute(
            """
            select order_no, year_folder, owner_folder, workbook_name, sheet_name, workbook_path,
                   outer_box_row, outer_box_spec, outer_box_qty, dimension_for_epl, c_n_for_epl, source
            from flow_content_index
            where upper(order_no)=upper(?)
            limit 1
            """,
            (order_no,),
        )
        row = cur.fetchone()
        if row is None:
            mapping = lookup_mapping(order_no)
            if mapping is None:
                return {
                    "ok": False,
                    "order_no": order_no,
                    "content_db": str(db),
                    "error": "order not found",
                }

            outer_box = read_outer_box_from_workbook(mapping["workbook_path"], mapping["sheet_name"])
            result = {
                "ok": outer_box is not None,
                "order_no": order_no,
                "content_db": str(db),
                "mapping_db": str(find_mapping_db()) if find_mapping_db() else None,
                "mapping": mapping,
                "fallback_used": True,
                "fallback_reason": "content index missing; resolved via order_mapping + workbook sheet",
            }
            if outer_box is None:
                result["error"] = "mapping found but outer box not found in workbook"
                return result

            result.update(outer_box)
            result["epl_fill_suggestion"] = {
                "c_n": outer_box.get("c_n_for_epl"),
                "dimension": outer_box.get("dimension_for_epl"),
                "carton_count": outer_box.get("outer_box_qty"),
                "packaging_source": "订单映射表回退读取流转单",
            }
            return result

        result = dict(row)
        result["ok"] = True
        result["content_db"] = str(db)
        result["fallback_used"] = False
        result["epl_fill_suggestion"] = {
            "c_n": result.get("c_n_for_epl"),
            "dimension": result.get("dimension_for_epl"),
            "carton_count": result.get("outer_box_qty"),
            "packaging_source": "流转单内容索引",
        }
        return result
    finally:
        conn.close()


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: query_flow_order.py <XS order no or text>", file=sys.stderr)
        return 2

    query = " ".join(sys.argv[1:])
    order_no = extract_order_no(query)
    if not order_no:
        print(json.dumps({"ok": False, "error": "XS order number not found", "query": query}, ensure_ascii=False, indent=2))
        return 1

    result = query_order(order_no)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
