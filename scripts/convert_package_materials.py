#!/usr/bin/env python3
"""Convert package material workbooks into a fast-query SQLite database."""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sqlite3
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_SQLITE_NAME = "包装资料.sqlite"


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
        force=True,
    )


def cleanup_abnormal_sqlite_files(output_dir: Path) -> None:
    for path in sorted(output_dir.glob("*.sqlite")):
        if path.name == DEFAULT_SQLITE_NAME:
            continue
        try:
            size = path.stat().st_size
        except OSError as exc:
            logging.warning("sqlite cleanup stat failed: path=%s error=%r", path, exc)
            continue
        if size > 0:
            continue
        try:
            path.unlink()
            logging.warning("sqlite cleanup removed empty unknown file: %s", path)
        except OSError as exc:
            logging.warning("sqlite cleanup unlink failed: path=%s error=%r", path, exc)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-root", required=True)
    parser.add_argument("--output-dir", required=True)
    return parser.parse_args()


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_float(value: object) -> float | None:
    text = cell_text(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def sheet_records(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, object]] = []
    try:
        for sheet in workbook.worksheets:
            if not sheet.max_row or not sheet.max_column:
                continue
            headers = [cell_text(v) for v in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            for row_idx, raw_row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row: dict[str, object] = {
                    "source_file": path.name,
                    "source_sheet": sheet.title,
                    "source_row": row_idx,
                }
                original: dict[str, str] = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = raw_row[idx] if idx < len(raw_row) else None
                    original[header] = cell_text(value)
                    row[header] = to_float(value) if isinstance(value, (int, float)) else cell_text(value)
                if any(v for k, v in row.items() if k not in {"source_file", "source_sheet", "source_row"}):
                    row["original_json"] = json.dumps(original, ensure_ascii=False)
                    rows.append(row)
    finally:
        workbook.close()
    return rows


def load_source_rows(source_root: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for path in sorted(source_root.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            rows.extend(sheet_records(path))
        except Exception as exc:
            logging.warning("failed to load %s: %r", path, exc)
    return rows


def write_sqlite(path: Path, rows: list[dict[str, object]]) -> None:
    columns = sorted({key for row in rows for key in row.keys()})
    if "original_json" not in columns:
        columns.append("original_json")
    if "source_row" not in columns:
        columns.append("source_row")

    with tempfile.NamedTemporaryFile(
        prefix="package-materials-",
        suffix=".sqlite",
        dir=path.parent,
        delete=False,
    ) as temp_file:
        temp_path = Path(temp_file.name)

    conn = sqlite3.connect(temp_path)
    try:
        schema = ["id INTEGER PRIMARY KEY AUTOINCREMENT"]
        for column in columns:
            schema.append(f'"{column}" {"INTEGER" if column == "source_row" else "TEXT"}')
        conn.execute(f'CREATE TABLE package_materials ({", ".join(schema)})')

        placeholders = ", ".join("?" for _ in columns)
        quoted_columns = ", ".join(f'"{column}"' for column in columns)
        conn.executemany(
            f"INSERT INTO package_materials ({quoted_columns}) VALUES ({placeholders})",
            ([row.get(column) for column in columns] for row in rows),
        )
        conn.execute("CREATE INDEX idx_package_materials_source_file ON package_materials (source_file)")
        conn.execute("CREATE INDEX idx_package_materials_source_sheet ON package_materials (source_sheet)")
        conn.execute("CREATE INDEX idx_package_materials_source_row ON package_materials (source_row)")
        conn.execute(
            """
            CREATE VIEW package_materials_lookup AS
            SELECT id, source_file, source_sheet, source_row, original_json
            FROM package_materials
            """
        )
        conn.commit()
    finally:
        conn.close()

    if path.exists():
        path.unlink()
    temp_path.rename(path)
    os.chmod(path, 0o666)


def main() -> int:
    configure_logging()
    args = parse_args()
    source_root = Path(args.source_root)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    cleanup_abnormal_sqlite_files(output_dir)

    if not source_root.exists():
        raise FileNotFoundError(source_root)

    rows = load_source_rows(source_root)
    sqlite_path = output_dir / DEFAULT_SQLITE_NAME
    write_sqlite(sqlite_path, rows)

    logging.info(
        json.dumps(
            {
                "generated_at": datetime.now().isoformat(timespec="seconds"),
                "source_root": str(source_root),
                "row_count": len(rows),
                "sqlite": str(sqlite_path),
                "sqlite_table": "package_materials",
                "sqlite_view": "package_materials_lookup",
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
